import math
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from lib.portfolio.construction import (
    compute_model_vol_42d,
    compute_portfolio_returns,
    load_portfolio_allocations,
    master_matrix_to_model_returns,
    risk_alloc_to_weights,
)


class PortfolioConstructionTests(unittest.TestCase):
    def test_master_matrix_to_model_returns_extracts_model_ids(self):
        master_matrix = pd.DataFrame(
            {
                "Date": pd.to_datetime(["2026-01-01", "2026-01-02"]),
                "1 - Model A": [0.01, 0.02],
                "2 - Model B": [-0.01, -0.02],
            }
        )
        loaded = master_matrix_to_model_returns(master_matrix)
        self.assertEqual(sorted(loaded["model_id"].unique().tolist()), [1, 2])
        self.assertEqual(len(loaded), 4)

    def test_compute_model_vol_42d_returns_nan_when_insufficient_history(self):
        model_returns_df = pd.DataFrame(
            {
                "date": pd.date_range("2026-01-01", periods=10, freq="D"),
                "model_id": [1] * 10,
                "model_return": [0.01] * 10,
            }
        )
        result = compute_model_vol_42d(model_returns_df)
        self.assertTrue(math.isnan(result.iloc[0]["vol_42d"]))

    def test_risk_alloc_to_weights_excludes_missing_vols_from_normalization(self):
        allocations_df = pd.DataFrame(
            {
                "portfolio_id": ["10014", "10014", "10014"],
                "model_id": [1, 2, 3],
                "risk_alloc": [0.5, 0.3, 0.2],
            }
        )
        model_vol_df = pd.DataFrame(
            {
                "model_id": [1, 2, 3],
                "vol_42d": [0.10, 0.20, float("nan")],
            }
        )
        weights = risk_alloc_to_weights(allocations_df, model_vol_df).set_index("model_id")
        self.assertAlmostEqual(weights.loc[1, "calc_risk_alloc_adjust"], 0.5)
        self.assertAlmostEqual(weights.loc[2, "calc_risk_alloc_adjust"], 0.3)
        self.assertAlmostEqual(weights.loc[3, "calc_risk_alloc_adjust"], 0.2)
        self.assertAlmostEqual(weights.loc[1, "calc_portfolio_wavg_vol"], 0.11)
        self.assertAlmostEqual(weights.loc[2, "calc_portfolio_wavg_vol"], 0.11)
        self.assertAlmostEqual(weights.loc[1, "calc_risk_weight_final"], 0.7692307692307692)
        self.assertAlmostEqual(weights.loc[2, "calc_risk_weight_final"], 0.23076923076923075)
        self.assertTrue(math.isnan(weights.loc[3, "calc_risk_weight_final"]))

    def test_compute_portfolio_returns_aligns_by_date(self):
        model_returns_df = pd.DataFrame(
            {
                "date": pd.to_datetime(["2026-01-01", "2026-01-01", "2026-01-02"]),
                "model_id": [1, 2, 1],
                "model_return": [0.01, 0.02, -0.03],
            }
        )
        weights_df = pd.DataFrame(
            {
                "portfolio_id": ["10014", "10014"],
                "model_id": [1, 2],
                "risk_alloc": [0.5, 0.5],
                "vol_42d": [0.1, 0.2],
                "calc_risk_weight_final": [0.6, 0.4],
            }
        )
        portfolio_returns = compute_portfolio_returns(model_returns_df, weights_df)
        self.assertAlmostEqual(portfolio_returns.iloc[0]["portfolio_return"], 0.014)
        self.assertAlmostEqual(portfolio_returns.iloc[1]["portfolio_return"], -0.018)

    def test_load_portfolio_allocations_filters_workbook_portfolio(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "Portfolio_Allocations.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "INDEX"
            ws.append(["portfolio_name", "portfolio_id", "sheet_name"])
            ws.append(["A", "10014", "A_10014"])
            ws.append(["B", "10015", "B_10015"])

            sheet_a = wb.create_sheet("A_10014")
            sheet_a.append(["portfolio_id", "model_id", "risk_alloc"])
            sheet_a.append(["10014", 1, 0.6])

            sheet_b = wb.create_sheet("B_10015")
            sheet_b.append(["portfolio_id", "model_id", "risk_alloc"])
            sheet_b.append(["10015", 2, 0.4])
            wb.save(path)

            loaded = load_portfolio_allocations(str(path), portfolio_id="10014")
            self.assertEqual(loaded["portfolio_id"].unique().tolist(), ["10014"])
            self.assertEqual(loaded["model_id"].tolist(), [1])


if __name__ == "__main__":
    unittest.main()
