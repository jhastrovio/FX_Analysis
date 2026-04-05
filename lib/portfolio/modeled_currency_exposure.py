"""
Helpers for single-date modeled currency exposure detail analysis.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from lib.portfolio.construction import compute_model_vol_42d, risk_alloc_to_weights


NON_CURRENCY_COLUMNS = {"date", "Category", "ID", "return", "return_ex_carry", "SPX", "US10Y"}


def _normalize_columns(df: pd.DataFrame, mapping: dict[str, list[str]], name: str) -> pd.DataFrame:
    result = df.copy()
    for target, candidates in mapping.items():
        for candidate in candidates:
            if candidate in result.columns:
                if candidate != target:
                    result = result.rename(columns={candidate: target})
                break
        else:
            raise ValueError(
                f"{name} missing required column '{target}'. Available columns: {list(result.columns)}"
            )
    return result


def load_allocations_with_strategy(path: str, portfolio_id: str) -> pd.DataFrame:
    """Load portfolio allocations with strategy metadata."""
    source_path = Path(path)
    suffix = source_path.suffix.lower()

    if suffix == ".csv":
        df = pd.read_csv(source_path)
        df = _normalize_columns(
            df,
            {
                "portfolio_id": ["portfolio_id", "Portfolio_ID", "PORTFOLIO_ID", "portfolio"],
                "model_id": ["model_id", "Model_ID", "MODEL_ID", "ID"],
                "strategy_name": ["strategy_name", "Strategy_Name", "STRATEGY_NAME", "Name"],
                "risk_alloc": ["risk_alloc", "Risk_Alloc", "RISK_ALLOC", "weight", "Weight"],
            },
            "Allocations CSV",
        )
    elif suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(source_path, data_only=True, read_only=True)
        if "INDEX" not in wb.sheetnames:
            raise ValueError("Allocations workbook missing required INDEX sheet")

        index_rows = list(wb["INDEX"].iter_rows(values_only=True))
        if not index_rows:
            raise ValueError("Allocations workbook INDEX sheet is empty")

        index_headers = [str(x).strip() if x is not None else "" for x in index_rows[0]]
        required_index_headers = {"portfolio_name", "portfolio_id", "sheet_name"}
        if not required_index_headers.issubset(set(index_headers)):
            raise ValueError(
                f"Allocations workbook INDEX sheet missing headers {required_index_headers}. "
                f"Found: {index_headers}"
            )

        idx_map = {name: index_headers.index(name) for name in required_index_headers}
        sheet_name: Optional[str] = None
        for row in index_rows[1:]:
            if row is None or row[idx_map["portfolio_id"]] is None:
                continue
            if str(row[idx_map["portfolio_id"]]) == str(portfolio_id):
                sheet_name = str(row[idx_map["sheet_name"]])
                break

        if sheet_name is None:
            raise ValueError(f"No allocation rows found for portfolio_id={portfolio_id}")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"INDEX references missing sheet: {sheet_name}")

        sheet_rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not sheet_rows:
            raise ValueError(f"Allocation sheet {sheet_name} is empty")

        headers = [str(x).strip() if x is not None else "" for x in sheet_rows[0]]
        required_headers = ["portfolio_id", "model_id", "strategy_name", "risk_alloc"]
        if not set(required_headers).issubset(set(headers)):
            raise ValueError(
                f"Sheet {sheet_name} missing required headers {required_headers}. Found: {headers}"
            )
        df = pd.DataFrame(sheet_rows[1:], columns=headers)
        df = df[[h for h in headers if h]].dropna(how="all")
    else:
        raise ValueError(f"Unsupported allocations file type: {suffix}")

    filtered = df[df["portfolio_id"].astype(str) == str(portfolio_id)].copy()
    if filtered.empty:
        raise ValueError(f"No allocation rows found for portfolio_id={portfolio_id}")

    filtered["portfolio_id"] = filtered["portfolio_id"].astype(str)
    filtered["model_id"] = pd.to_numeric(filtered["model_id"], errors="raise").astype(int)
    filtered["risk_alloc"] = pd.to_numeric(filtered["risk_alloc"], errors="coerce")
    filtered["strategy_name"] = filtered["strategy_name"].astype(str)
    return filtered[["portfolio_id", "model_id", "strategy_name", "risk_alloc"]].sort_values("model_id")


def discover_latest_model_files(model_ids: list[int], model_data_dir: str) -> dict[int, Path]:
    """Locate the latest available model file for each selected model."""
    base_dir = Path(model_data_dir)
    file_map: dict[int, Path] = {}
    missing: list[int] = []

    for model_id in sorted(set(model_ids)):
        matches = sorted(base_dir.glob(f"{model_id}_*.csv"))
        if not matches:
            missing.append(model_id)
            continue
        file_map[model_id] = matches[-1]

    if missing:
        raise ValueError(f"Missing model files for model_ids={missing} in {base_dir}")

    return file_map


def load_model_histories(model_files: dict[int, Path]) -> dict[int, pd.DataFrame]:
    """Load model history CSVs keyed by model_id."""
    histories: dict[int, pd.DataFrame] = {}
    for model_id, path in model_files.items():
        df = pd.read_csv(path)
        df = _normalize_columns(
            df,
            {
                "date": ["date", "Date", "DATE"],
                "model_id": ["model_id", "Model_ID", "MODEL_ID", "ID"],
                "model_return": ["model_return", "return", "daily_return"],
            },
            f"Model file {path.name}",
        )
        df["date"] = pd.to_datetime(df["date"])
        df["model_id"] = pd.to_numeric(df["model_id"], errors="raise").astype(int)
        df["model_return"] = pd.to_numeric(df["model_return"], errors="coerce")
        histories[model_id] = df
    return histories


def get_currency_columns(df: pd.DataFrame) -> list[str]:
    """Return the currency exposure columns from a model history file."""
    return [
        column
        for column in df.columns
        if column not in NON_CURRENCY_COLUMNS and len(str(column)) == 3 and str(column).isalpha() and str(column).isupper()
    ]


def resolve_target_date(model_histories: dict[int, pd.DataFrame], requested_date: Optional[str] = None) -> pd.Timestamp:
    """Resolve the target date from the common date intersection."""
    if not model_histories:
        raise ValueError("No model histories supplied")

    common_dates = None
    for df in model_histories.values():
        available = set(pd.to_datetime(df["date"]).dt.normalize())
        common_dates = available if common_dates is None else common_dates & available

    ordered = sorted(common_dates or [])
    if not ordered:
        raise ValueError("No common dates found across selected model files")

    if requested_date is None:
        return ordered[-1]

    requested = pd.to_datetime(requested_date).normalize()
    if requested in common_dates:
        return requested

    nearest = sorted(ordered, key=lambda dt: abs((dt - requested).days))[:5]
    nearest_str = ", ".join(dt.strftime("%Y-%m-%d") for dt in nearest)
    raise ValueError(
        f"Requested date {requested.strftime('%Y-%m-%d')} is not available across all selected models. "
        f"Nearest common dates: {nearest_str}"
    )


def build_model_returns_df(model_histories: dict[int, pd.DataFrame]) -> pd.DataFrame:
    """Assemble long-form returns needed for trailing volatility calculation."""
    frames = [
        df[["date", "model_id", "model_return"]].copy()
        for df in model_histories.values()
    ]
    return pd.concat(frames, ignore_index=True).sort_values(["model_id", "date"]).reset_index(drop=True)


def extract_modeled_position_snapshot(
    model_histories: dict[int, pd.DataFrame],
    target_date: pd.Timestamp,
) -> pd.DataFrame:
    """Extract long-form non-zero modeled currency positions for one date."""
    rows: list[pd.DataFrame] = []
    for model_id, df in model_histories.items():
        currencies = get_currency_columns(df)
        snapshot = df.loc[df["date"].dt.normalize() == target_date, ["date", "model_id", *currencies]].copy()
        if snapshot.empty:
            raise ValueError(f"Model {model_id} does not contain target date {target_date.strftime('%Y-%m-%d')}")
        if len(snapshot) > 1:
            snapshot = snapshot.tail(1)

        melted = snapshot.melt(
            id_vars=["date", "model_id"],
            value_vars=currencies,
            var_name="currency",
            value_name="model_position_pct",
        )
        melted["model_position_pct"] = pd.to_numeric(melted["model_position_pct"], errors="coerce")
        melted = melted[melted["model_position_pct"].notna() & (melted["model_position_pct"] != 0)].copy()
        rows.append(melted)

    if not rows:
        return pd.DataFrame(columns=["date", "model_id", "currency", "model_position_pct"])

    return pd.concat(rows, ignore_index=True).sort_values(["model_id", "currency"]).reset_index(drop=True)


def build_modeled_currency_exposure_detail(
    allocations_df: pd.DataFrame,
    model_histories: dict[int, pd.DataFrame],
    target_date: pd.Timestamp,
) -> pd.DataFrame:
    """Build the modeled currency exposure detail table for one date."""
    model_returns_df = build_model_returns_df(model_histories)
    model_vol_df = compute_model_vol_42d(model_returns_df)
    weights_df = risk_alloc_to_weights(allocations_df, model_vol_df).rename(
        columns={
            "calc_risk_alloc_adjust": "risk_alloc_norm",
            "calc_portfolio_wavg_vol": "portfolio_wavg_vol",
        }
    )
    weights_df["vol_gap"] = (
        weights_df["portfolio_wavg_vol"] - weights_df["vol_42d"]
    ) / weights_df["portfolio_wavg_vol"]
    weights_df["gap_scalar"] = (1.0 + weights_df["vol_gap"]).clip(lower=0.0)
    weights_df["raw_weight"] = weights_df["risk_alloc_norm"] * weights_df["gap_scalar"]
    weight_sum = weights_df["raw_weight"].sum(min_count=1)
    weights_df["raw_weight_sum"] = weight_sum
    weights_df["final_weight"] = weights_df["raw_weight"] / weights_df["raw_weight_sum"]

    positions_df = extract_modeled_position_snapshot(model_histories, target_date)
    detail_df = positions_df.merge(
        weights_df[
            [
                "portfolio_id",
                "model_id",
                "strategy_name",
                "risk_alloc",
                "risk_alloc_norm",
                "vol_42d",
                "portfolio_wavg_vol",
                "vol_gap",
                "gap_scalar",
                "raw_weight",
                "final_weight",
                "raw_weight_sum",
            ]
        ],
        on="model_id",
        how="left",
    )
    detail_df = detail_df[detail_df["final_weight"].notna()].copy()
    detail_df["modeled_exposure_usd_equiv"] = detail_df["final_weight"] * detail_df["model_position_pct"]

    ordered_columns = [
        "date",
        "portfolio_id",
        "model_id",
        "strategy_name",
        "risk_alloc",
        "risk_alloc_norm",
        "vol_42d",
        "portfolio_wavg_vol",
        "vol_gap",
        "gap_scalar",
        "raw_weight",
        "final_weight",
        "raw_weight_sum",
        "currency",
        "model_position_pct",
        "modeled_exposure_usd_equiv",
    ]
    detail_df = detail_df[ordered_columns].sort_values(["model_id", "currency"]).reset_index(drop=True)

    total_rows = (
        detail_df.groupby(["date", "portfolio_id", "currency"], as_index=False)
        .agg(
            modeled_exposure_usd_equiv=("modeled_exposure_usd_equiv", "sum"),
        )
    )
    total_rows["model_id"] = 0
    total_rows["strategy_name"] = "total"
    total_rows["risk_alloc"] = 0.0
    total_rows["risk_alloc_norm"] = 0.0
    total_rows["vol_42d"] = 0.0
    total_rows["portfolio_wavg_vol"] = detail_df["portfolio_wavg_vol"].iloc[0] if not detail_df.empty else 0.0
    total_rows["vol_gap"] = 0.0
    total_rows["gap_scalar"] = 0.0
    total_rows["raw_weight"] = 0.0
    total_rows["final_weight"] = 0.0
    total_rows["raw_weight_sum"] = weight_sum if pd.notna(weight_sum) else 0.0
    total_rows["model_position_pct"] = 0.0
    total_rows = total_rows[ordered_columns]

    combined = pd.concat([detail_df, total_rows], ignore_index=True)
    return combined.sort_values(["model_id", "currency"]).reset_index(drop=True)
