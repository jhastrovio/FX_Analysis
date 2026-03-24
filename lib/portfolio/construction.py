"""
Lightweight helpers for on-demand portfolio construction analytics.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from lib.risk.volatility import trailing_vol


def _normalize_columns(df: pd.DataFrame, mapping: dict[str, list[str]], name: str) -> pd.DataFrame:
    """Rename plausible source columns to the expected internal schema."""
    result = df.copy()
    for target, candidates in mapping.items():
        for candidate in candidates:
            if candidate in result.columns:
                if candidate != target:
                    result = result.rename(columns={candidate: target})
                break
        else:
            raise ValueError(
                f"{name} missing required column '{target}'. Available columns: {list(result.columns)}"
            )
    return result


def load_portfolio_allocations(path: str, portfolio_id: Optional[str] = None) -> pd.DataFrame:
    """Load portfolio allocations from CSV or the static workbook."""
    suffix = Path(path).suffix.lower()

    if suffix == ".csv":
        df = pd.read_csv(path)
        df = _normalize_columns(
            df,
            {
                "portfolio_id": ["portfolio_id", "Portfolio_ID", "PORTFOLIO_ID", "portfolio"],
                "model_id": ["model_id", "Model_ID", "MODEL_ID", "ID"],
                "risk_alloc": ["risk_alloc", "Risk_Alloc", "RISK_ALLOC", "weight", "Weight"],
            },
            "Allocations CSV",
        )
    elif suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        if "INDEX" not in wb.sheetnames:
            raise ValueError("Allocations workbook missing required INDEX sheet")

        index_rows = list(wb["INDEX"].iter_rows(values_only=True))
        if not index_rows:
            raise ValueError("Allocations workbook INDEX sheet is empty")

        index_headers = [str(x).strip() if x is not None else "" for x in index_rows[0]]
        required_index_headers = {"portfolio_name", "portfolio_id", "sheet_name"}
        if not required_index_headers.issubset(set(index_headers)):
            raise ValueError(
                f"Allocations workbook INDEX sheet missing headers {required_index_headers}. "
                f"Found: {index_headers}"
            )

        idx_map = {name: index_headers.index(name) for name in required_index_headers}
        frames = []
        for row in index_rows[1:]:
            if row is None or row[idx_map["portfolio_id"]] is None:
                continue
            sheet_name = row[idx_map["sheet_name"]]
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"INDEX references missing sheet: {sheet_name}")
            sheet_rows = list(wb[sheet_name].iter_rows(values_only=True))
            if not sheet_rows:
                continue

            headers = [str(x).strip() if x is not None else "" for x in sheet_rows[0]]
            required_headers = ["portfolio_id", "model_id", "risk_alloc"]
            if not set(required_headers).issubset(set(headers)):
                raise ValueError(
                    f"Sheet {sheet_name} missing required headers {required_headers}. "
                    f"Found: {headers}"
                )
            sheet_df = pd.DataFrame(sheet_rows[1:], columns=headers)
            sheet_df = sheet_df[[h for h in headers if h]]
            sheet_df = sheet_df.dropna(how="all")
            frames.append(sheet_df[required_headers])

        if not frames:
            raise ValueError("Allocations workbook contains no allocation rows")

        df = pd.concat(frames, ignore_index=True)
    else:
        raise ValueError(f"Unsupported allocations file type: {suffix}")

    df["portfolio_id"] = df["portfolio_id"].astype(str)
    if portfolio_id is None:
        unique_portfolios = sorted(df["portfolio_id"].dropna().unique())
        if len(unique_portfolios) != 1:
            raise ValueError(
                "Allocations source contains multiple portfolio_ids. Pass --portfolio-id explicitly."
            )
        portfolio_id = unique_portfolios[0]

    filtered = df[df["portfolio_id"] == str(portfolio_id)].copy()
    if filtered.empty:
        raise ValueError(f"No allocation rows found for portfolio_id={portfolio_id}")

    filtered["model_id"] = pd.to_numeric(filtered["model_id"], errors="raise").astype(int)
    filtered["risk_alloc"] = pd.to_numeric(filtered["risk_alloc"], errors="coerce")
    return filtered[["portfolio_id", "model_id", "risk_alloc"]]


def master_matrix_to_model_returns(master_matrix: pd.DataFrame) -> pd.DataFrame:
    """Convert the existing wide master matrix into long-form model returns."""
    if master_matrix.index.name != "Date":
        matrix = master_matrix.copy()
        if "Date" in matrix.columns:
            matrix["Date"] = pd.to_datetime(matrix["Date"])
            matrix = matrix.set_index("Date")
        else:
            matrix.index = pd.to_datetime(matrix.index)
    else:
        matrix = master_matrix.copy()
        matrix.index = pd.to_datetime(matrix.index)

    model_columns = [col for col in matrix.columns if " - " in str(col)]
    long_df = matrix[model_columns].reset_index().melt(
        id_vars=["Date"],
        value_vars=model_columns,
        var_name="model_column",
        value_name="model_return",
    )
    long_df["model_id"] = long_df["model_column"].str.split(" - ", n=1).str[0].astype(int)
    long_df = long_df.rename(columns={"Date": "date"})
    return long_df[["date", "model_id", "model_return"]]


def load_model_returns(path: str) -> pd.DataFrame:
    """Load model returns from either long or master-matrix-style wide CSV."""
    df = pd.read_csv(path)
    if "Date" in df.columns and any(" - " in str(col) for col in df.columns):
        df["Date"] = pd.to_datetime(df["Date"])
        return master_matrix_to_model_returns(df)

    df = _normalize_columns(
        df,
        {
            "date": ["date", "Date", "DATE"],
            "model_id": ["model_id", "Model_ID", "MODEL_ID", "ID"],
            "model_return": ["model_return", "return", "daily_return"],
        },
        "Model returns CSV",
    )
    df["date"] = pd.to_datetime(df["date"])
    df["model_id"] = pd.to_numeric(df["model_id"], errors="raise").astype(int)
    df["model_return"] = pd.to_numeric(df["model_return"], errors="coerce")
    return df[["date", "model_id", "model_return"]]


def compute_model_vol_42d(model_returns_df: pd.DataFrame) -> pd.DataFrame:
    """Compute 42-day annualized vol for each model."""
    result_rows = []
    for model_id, group in model_returns_df.groupby("model_id", sort=True):
        vol_42d = trailing_vol(group.sort_values("date")["model_return"], window=42)
        result_rows.append({"model_id": model_id, "vol_42d": vol_42d})
    return pd.DataFrame(result_rows)


def risk_alloc_to_weights(
    allocations_df: pd.DataFrame,
    model_vol_df: pd.DataFrame,
) -> pd.DataFrame:
    """Convert risk allocations into normalized, vol-adjusted portfolio weights."""
    weights = allocations_df.merge(model_vol_df, on="model_id", how="left")
    valid = weights["vol_42d"].notna() & np.isfinite(weights["vol_42d"]) & (weights["vol_42d"] > 0)
    risk_alloc_sum = weights.loc[weights["risk_alloc"].notna(), "risk_alloc"].sum()
    if risk_alloc_sum > 0:
        weights["calc_risk_alloc_adjust"] = weights["risk_alloc"] / risk_alloc_sum
    else:
        weights["calc_risk_alloc_adjust"] = np.nan

    portfolio_wavg_vol = (
        weights.loc[valid, "calc_risk_alloc_adjust"] * weights.loc[valid, "vol_42d"]
    ).sum(min_count=1)
    weights["calc_portfolio_wavg_vol"] = portfolio_wavg_vol
    weights["calc_prelim_risk_weight"] = np.where(
        valid,
        weights["calc_risk_alloc_adjust"] * portfolio_wavg_vol / weights["vol_42d"],
        np.nan,
    )

    prelim_sum = weights.loc[weights["calc_prelim_risk_weight"].notna(), "calc_prelim_risk_weight"].sum()
    if prelim_sum > 0:
        weights["calc_risk_weight_final"] = weights["calc_prelim_risk_weight"] / prelim_sum
    else:
        weights["calc_risk_weight_final"] = np.nan

    preferred_cols = [
        column
        for column in [
            "portfolio_id",
            "model_id",
            "strategy_name",
            "risk_alloc",
            "calc_risk_alloc_adjust",
            "vol_42d",
            "calc_portfolio_wavg_vol",
            "calc_prelim_risk_weight",
            "calc_risk_weight_final",
        ]
        if column in weights.columns
    ]
    return weights[preferred_cols].sort_values("model_id")


def compute_portfolio_returns(
    model_returns_df: pd.DataFrame,
    weights_df: pd.DataFrame,
) -> pd.DataFrame:
    """Construct portfolio daily returns from model returns and derived weights."""
    weight_column = "calc_risk_weight_final" if "calc_risk_weight_final" in weights_df.columns else "weight"
    weighted = model_returns_df.merge(weights_df[["model_id", weight_column]], on="model_id", how="inner")
    weighted = weighted[weighted[weight_column].notna()].copy()
    weighted["weighted_return"] = weighted[weight_column] * weighted["model_return"]

    portfolio_returns = (
        weighted.groupby("date", as_index=False)
        .agg(portfolio_return=("weighted_return", lambda s: s.sum(min_count=1)))
        .sort_values("date")
        .reset_index(drop=True)
    )
    return portfolio_returns
