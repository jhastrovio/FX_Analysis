#!/usr/bin/env python3
"""
Analytics-only review of a live portfolio subset against the all-model universe.

This module reads local workbook snapshots and writes an ephemeral summary workbook.
It does not modify source workbooks or produce canonical datasets.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
TEXT_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
SOURCE_ALL_KEY = "allmodels"
SOURCE_PORTFOLIO_KEY = "portfolio"


@dataclass
class PortfolioReview:
    portfolio_rows: List[Dict[str, object]]
    all_model_rows: List[Dict[str, object]]
    category_summary: List[Dict[str, object]]
    family_summary: List[Dict[str, object]]
    top_missing_rows: List[Dict[str, object]]
    rebalance_rows: List[Dict[str, object]]
    summary_rows: List[List[object]]


def _clean_header(value: object) -> str:
    return str(value).replace("\xa0", " ").strip()


def _sheet_rows(path: Path) -> List[Dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [_clean_header(cell.value) for cell in sheet[1]]
    rows: List[Dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def _weighted_average(rows: Iterable[Dict[str, object]], value_key: str, weight_key: str) -> float | None:
    total_weight = 0.0
    weighted_total = 0.0
    for row in rows:
        value = row.get(value_key)
        weight = row.get(weight_key)
        if isinstance(value, (int, float)) and isinstance(weight, (int, float)):
            total_weight += float(weight)
            weighted_total += float(value) * float(weight)
    if total_weight == 0:
        return None
    return weighted_total / total_weight


def _mean(rows: Iterable[Dict[str, object]], key: str) -> float | None:
    values = [float(row[key]) for row in rows if isinstance(row.get(key), (int, float))]
    if not values:
        return None
    return sum(values) / len(values)


def _metric_ranks(rows: List[Dict[str, object]], metric_key: str) -> Dict[int, int]:
    ranked_rows = sorted(
        [row for row in rows if isinstance(row.get(metric_key), (int, float))],
        key=lambda row: float(row[metric_key]),
        reverse=True,
    )
    return {int(row["ID"]): index for index, row in enumerate(ranked_rows, start=1)}


def _rank_score(rank: int, total_count: int) -> float:
    if total_count <= 1:
        return 1.0
    return (total_count - rank) / (total_count - 1)


def _summarize_group(rows: List[Dict[str, object]], group_key: str) -> List[Dict[str, object]]:
    grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[str(row[group_key])].append(row)

    summary: List[Dict[str, object]] = []
    for group_name, group_rows in grouped.items():
        weight_sum = sum(float(row["Risk Alloc"]) for row in group_rows if isinstance(row.get("Risk Alloc"), (int, float)))
        summary.append(
            {
                group_key: group_name,
                "Model Count": len(group_rows),
                "Total Weight": weight_sum,
                "Weighted 1D Ret": _weighted_average(group_rows, "1D Ret", "Risk Alloc"),
                "Weighted 2D Ret": _weighted_average(group_rows, "2D Ret", "Risk Alloc"),
                "Weighted 5D Ret": _weighted_average(group_rows, "5D Ret", "Risk Alloc"),
                "Avg SR 2M": _mean(group_rows, "SR 2M"),
                "Avg SR 1Y": _mean(group_rows, "SR 1Y"),
                "Avg Composite Rank": _mean(group_rows, "Composite Rank 33/33/33"),
            }
        )

    return sorted(summary, key=lambda row: (-float(row["Total Weight"]), str(row[group_key])))


def build_review(allmodels_path: Path, portfolio_path: Path, missing_limit: int = 20) -> PortfolioReview:
    all_rows = _sheet_rows(allmodels_path)
    portfolio_rows = _sheet_rows(portfolio_path)

    all_by_id = {int(row["ID"]): row for row in all_rows if row.get("ID") is not None}
    sr_2m_ranks = _metric_ranks(all_rows, "SR 2M*")
    sr_1y_ranks = _metric_ranks(all_rows, "SR 1Y*")
    sr_2k_ranks = _metric_ranks(all_rows, "SR 2K*")
    universe_count = len(all_by_id)

    ranked_all: List[Dict[str, object]] = []
    for row in all_rows:
        model_id = int(row["ID"])
        rank_2m = sr_2m_ranks[model_id]
        rank_1y = sr_1y_ranks[model_id]
        rank_2k = sr_2k_ranks[model_id]
        composite_score = (
            _rank_score(rank_2m, universe_count)
            + _rank_score(rank_1y, universe_count)
            + _rank_score(rank_2k, universe_count)
        ) / 3.0
        ranked_all.append(
            {
                **row,
                "Rank SR 2M": rank_2m,
                "Rank SR 1Y": rank_1y,
                "Rank SR 2K": rank_2k,
                "Composite Score 33/33/33": composite_score,
            }
        )

    ranked_all.sort(key=lambda row: float(row["Composite Score 33/33/33"]), reverse=True)
    composite_ranks = {int(row["ID"]): index for index, row in enumerate(ranked_all, start=1)}
    all_model_rows: List[Dict[str, object]] = []
    for row in ranked_all:
        model_id = int(row["ID"])
        all_model_rows.append(
            {
                "Composite Rank 33/33/33": composite_ranks[model_id],
                "Composite Score 33/33/33": row["Composite Score 33/33/33"],
                "Rank SR 2M": row["Rank SR 2M"],
                "Rank SR 1Y": row["Rank SR 1Y"],
                "Rank SR 2K": row["Rank SR 2K"],
                "ID": model_id,
                "Name": row["Name"],
                "Category": row["CATEGORY"],
                "Family": row["FAMILY"],
                "SR 2M": row["SR 2M*"],
                "SR 1Y": row["SR 1Y*"],
                "SR 2K": row["SR 2K*"],
                "1D Ret": row["1D Ret"],
                "5D Ret": row["5D Ret"],
                "SPX 2K": row["SPX 2K**"],
                "SPX 2M": row["SPX 2M**"],
                "10Y 2K": row["10Y 2K**"],
                "10Y 2M": row["10Y 2M**"],
                "Held In Portfolio": "Yes" if model_id in {int(r['ID']) for r in portfolio_rows if r.get('ID') is not None} else "No",
                "Source": str(allmodels_path),
            }
        )

    enriched_portfolio: List[Dict[str, object]] = []
    held_ids = set()
    for portfolio_row in portfolio_rows:
        model_id = int(portfolio_row["ID"])
        held_ids.add(model_id)
        master_row = all_by_id[model_id]
        category, family = str(portfolio_row["CAT. / FAMILY"]).split(" / ", 1)
        enriched_portfolio.append(
            {
                "ID": model_id,
                "Strategy Name": portfolio_row["Strategy Name"],
                "Category": category,
                "Family": family,
                "Risk Alloc": portfolio_row["Risk Alloc"],
                "1D Ret": portfolio_row["1D Ret"],
                "2D Ret": portfolio_row["2D Ret"],
                "5D Ret": portfolio_row["5D Ret"],
                "SR 2M": master_row["SR 2M*"],
                "SR 1Y": master_row["SR 1Y*"],
                "SR 2K": master_row["SR 2K*"],
                "Rank SR 2M": sr_2m_ranks.get(model_id),
                "Rank SR 1Y": sr_1y_ranks.get(model_id),
                "Rank SR 2K": sr_2k_ranks.get(model_id),
                "Composite Score 33/33/33": next(row["Composite Score 33/33/33"] for row in ranked_all if int(row["ID"]) == model_id),
                "Composite Rank 33/33/33": composite_ranks.get(model_id),
                "SPX 2K": master_row["SPX 2K**"],
                "SPX 2M": master_row["SPX 2M**"],
                "10Y 2K": master_row["10Y 2K**"],
                "10Y 2M": master_row["10Y 2M**"],
                "Allmodels Name": master_row["Name"],
                "Source": str(portfolio_path),
            }
        )

    enriched_portfolio.sort(
        key=lambda row: (
            float(row["Composite Rank 33/33/33"]) if isinstance(row.get("Composite Rank 33/33/33"), (int, float)) else 10**9,
            -float(row["Risk Alloc"]) if isinstance(row.get("Risk Alloc"), (int, float)) else 0.0,
        )
    )

    missing_rows: List[Dict[str, object]] = []
    for ranked_row in ranked_all:
        model_id = int(ranked_row["ID"])
        if model_id in held_ids:
            continue
        missing_rows.append(
            {
                "Composite Rank 33/33/33": composite_ranks[model_id],
                "Rank SR 2M": sr_2m_ranks[model_id],
                "Rank SR 1Y": sr_1y_ranks[model_id],
                "Rank SR 2K": sr_2k_ranks[model_id],
                "ID": model_id,
                "Name": ranked_row["Name"],
                "Category": ranked_row["CATEGORY"],
                "Family": ranked_row["FAMILY"],
                "Composite Score 33/33/33": ranked_row["Composite Score 33/33/33"],
                "SR 2M": ranked_row["SR 2M*"],
                "SR 1Y": ranked_row["SR 1Y*"],
                "SR 2K": ranked_row["SR 2K*"],
                "1D Ret": ranked_row["1D Ret"],
                "5D Ret": ranked_row["5D Ret"],
                "Source": str(allmodels_path),
            }
        )
        if len(missing_rows) >= missing_limit:
            break

    unheld_by_category: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    unheld_by_family: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in all_model_rows:
        if row["Held In Portfolio"] == "Yes":
            continue
        unheld_by_category[str(row["Category"])].append(row)
        unheld_by_family[str(row["Family"])].append(row)

    rebalance_rows: List[Dict[str, object]] = []
    weaker_held = sorted(
        enriched_portfolio,
        key=lambda row: float(row["Composite Rank 33/33/33"]),
        reverse=True,
    )
    for held_row in weaker_held:
        family_candidates = unheld_by_family.get(str(held_row["Family"]), [])
        category_candidates = unheld_by_category.get(str(held_row["Category"]), [])
        family_upgrade = next(
            (
                candidate
                for candidate in family_candidates
                if float(candidate["Composite Rank 33/33/33"]) < float(held_row["Composite Rank 33/33/33"])
            ),
            None,
        )
        category_upgrade = next(
            (
                candidate
                for candidate in category_candidates
                if float(candidate["Composite Rank 33/33/33"]) < float(held_row["Composite Rank 33/33/33"])
            ),
            None,
        )
        if family_upgrade is None and category_upgrade is None:
            continue
        chosen_upgrade = family_upgrade or category_upgrade
        comparison_scope = "same family" if family_upgrade else "same category"
        rebalance_rows.append(
            {
                "Held ID": held_row["ID"],
                "Held Strategy": held_row["Strategy Name"],
                "Held Category": held_row["Category"],
                "Held Family": held_row["Family"],
                "Held Weight": held_row["Risk Alloc"],
                "Held Composite Rank": held_row["Composite Rank 33/33/33"],
                "Held Composite Score": held_row["Composite Score 33/33/33"],
                "Held SR 2M": held_row["SR 2M"],
                "Held SR 1Y": held_row["SR 1Y"],
                "Held SR 2K": held_row["SR 2K"],
                "Upgrade Scope": comparison_scope,
                "Candidate ID": chosen_upgrade["ID"],
                "Candidate Name": chosen_upgrade["Name"],
                "Candidate Category": chosen_upgrade["Category"],
                "Candidate Family": chosen_upgrade["Family"],
                "Candidate Composite Rank": chosen_upgrade["Composite Rank 33/33/33"],
                "Candidate Composite Score": chosen_upgrade["Composite Score 33/33/33"],
                "Candidate SR 2M": chosen_upgrade["SR 2M"],
                "Candidate SR 1Y": chosen_upgrade["SR 1Y"],
                "Candidate SR 2K": chosen_upgrade["SR 2K"],
                "Rank Improvement": float(held_row["Composite Rank 33/33/33"]) - float(chosen_upgrade["Composite Rank 33/33/33"]),
                "Source": str(allmodels_path),
            }
        )

    rebalance_rows.sort(key=lambda row: (-float(row["Rank Improvement"]), float(row["Held Composite Rank"])))

    top_10 = sum(1 for row in enriched_portfolio if isinstance(row["Composite Rank 33/33/33"], (int, float)) and row["Composite Rank 33/33/33"] <= 10)
    top_20 = sum(1 for row in enriched_portfolio if isinstance(row["Composite Rank 33/33/33"], (int, float)) and row["Composite Rank 33/33/33"] <= 20)
    top_30 = sum(1 for row in enriched_portfolio if isinstance(row["Composite Rank 33/33/33"], (int, float)) and row["Composite Rank 33/33/33"] <= 30)
    negative_sr_held = sum(1 for row in enriched_portfolio if isinstance(row["SR 2M"], (int, float)) and float(row["SR 2M"]) < 0)
    total_weight = sum(float(row["Risk Alloc"]) for row in enriched_portfolio if isinstance(row.get("Risk Alloc"), (int, float)))
    best_held = min(enriched_portfolio, key=lambda row: float(row["Composite Rank 33/33/33"]))
    worst_held = max(enriched_portfolio, key=lambda row: float(row["Composite Rank 33/33/33"]))

    summary_rows = [
        ["Metric", "Value"],
        ["Portfolio model count", len(enriched_portfolio)],
        ["Risk allocation sum", total_weight],
        ["Weighted 1D return", _weighted_average(enriched_portfolio, "1D Ret", "Risk Alloc")],
        ["Weighted 2D return", _weighted_average(enriched_portfolio, "2D Ret", "Risk Alloc")],
        ["Weighted 5D return", _weighted_average(enriched_portfolio, "5D Ret", "Risk Alloc")],
        ["Weighted avg SR 2M", _weighted_average(enriched_portfolio, "SR 2M", "Risk Alloc")],
        ["Weighted avg SR 1Y", _weighted_average(enriched_portfolio, "SR 1Y", "Risk Alloc")],
        ["Weighted avg SR 2K", _weighted_average(enriched_portfolio, "SR 2K", "Risk Alloc")],
        ["Average composite rank", _mean(enriched_portfolio, "Composite Rank 33/33/33")],
        ["Weighted average composite rank", _weighted_average(enriched_portfolio, "Composite Rank 33/33/33", "Risk Alloc")],
        ["Held models in top 10 composite", top_10],
        ["Held models in top 20 composite", top_20],
        ["Held models in top 30 composite", top_30],
        ["Held models with negative SR 2M", negative_sr_held],
        ["Best-held composite rank", f"{best_held['Composite Rank 33/33/33']} | {best_held['Strategy Name']}"],
        ["Worst-held composite rank", f"{worst_held['Composite Rank 33/33/33']} | {worst_held['Strategy Name']}"],
        [f"Source: {SOURCE_ALL_KEY}", str(allmodels_path)],
        [f"Source: {SOURCE_PORTFOLIO_KEY}", str(portfolio_path)],
    ]

    return PortfolioReview(
        portfolio_rows=enriched_portfolio,
        all_model_rows=all_model_rows,
        category_summary=_summarize_group(enriched_portfolio, "Category"),
        family_summary=_summarize_group(enriched_portfolio, "Family"),
        top_missing_rows=missing_rows,
        rebalance_rows=rebalance_rows,
        summary_rows=summary_rows,
    )


def _apply_header_style(cells) -> None:
    for cell in cells:
        cell.fill = HEADER_FILL
        cell.font = TEXT_FONT
        cell.alignment = HEADER_ALIGNMENT


def _autofit_sheet(sheet) -> None:
    widths: Dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 11), 36)


def _apply_numeric_formats(sheet, column_names: Dict[str, str]) -> None:
    header_map = {cell.value: cell.column for cell in sheet[1]}
    weight_columns = {"Risk Alloc", "Total Weight"}
    return_columns = {"1D Ret", "2D Ret", "5D Ret", "Weighted 1D Ret", "Weighted 2D Ret", "Weighted 5D Ret"}
    stat_columns = {
        "SR 2M",
        "SR 1Y",
        "SR 2K",
        "SPX 2K",
        "SPX 2M",
        "10Y 2K",
        "10Y 2M",
        "Avg SR 2M",
        "Avg SR 1Y",
        "Composite Score 33/33/33",
        "Held Composite Score",
        "Candidate Composite Score",
        "Held SR 2M",
        "Held SR 1Y",
        "Held SR 2K",
        "Candidate SR 2M",
        "Candidate SR 1Y",
        "Candidate SR 2K",
    }
    rank_columns = {
        "Avg Composite Rank",
        "Rank SR 2M",
        "Rank SR 1Y",
        "Rank SR 2K",
        "Composite Rank 33/33/33",
        "Held Composite Rank",
        "Candidate Composite Rank",
        "Rank Improvement",
    }

    for column_name, column_index in header_map.items():
        if column_name in weight_columns:
            fmt = "0.0%"
        elif column_name in return_columns or column_name in stat_columns:
            fmt = '0.00;[Red](0.00);-'
        elif column_name in rank_columns:
            fmt = '0.0'
        else:
            continue
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column_index).number_format = fmt


def _write_table(sheet, rows: List[Dict[str, object]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    _apply_header_style(sheet[1])
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    _apply_numeric_formats(sheet, {})
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autofit_sheet(sheet)


def write_review_workbook(review: PortfolioReview, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    for row in review.summary_rows:
        summary_sheet.append(row)
    _apply_header_style(summary_sheet[1])
    for row_index in range(2, summary_sheet.max_row + 1):
        label_cell = summary_sheet.cell(row=row_index, column=1)
        label_cell.fill = SECTION_FILL
        label_cell.font = Font(bold=True)
    summary_sheet["B3"].number_format = "0.0%"
    for cell_ref in ("B4", "B5", "B6", "B7", "B8", "B9"):
        summary_sheet[cell_ref].number_format = '0.00;[Red](0.00);-'
    for cell_ref in ("B10", "B11"):
        summary_sheet[cell_ref].number_format = "0.0"
    _autofit_sheet(summary_sheet)

    portfolio_sheet = workbook.create_sheet("Portfolio Models")
    _write_table(portfolio_sheet, review.portfolio_rows)

    all_models_sheet = workbook.create_sheet("All Models")
    _write_table(all_models_sheet, review.all_model_rows)

    category_sheet = workbook.create_sheet("Category Summary")
    _write_table(category_sheet, review.category_summary)

    family_sheet = workbook.create_sheet("Family Summary")
    _write_table(family_sheet, review.family_summary)

    missing_sheet = workbook.create_sheet("Top Ranked Missing")
    _write_table(missing_sheet, review.top_missing_rows)

    rebalance_sheet = workbook.create_sheet("Rebalance Candidates")
    _write_table(rebalance_sheet, review.rebalance_rows)

    workbook.save(output_path)
    return output_path
