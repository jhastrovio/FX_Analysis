"""
Reusable validation helpers for reproducing model summary metrics from model history CSVs.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TRADING_DAYS = 252
ROLLING_1Y_BUSINESS_DAYS = 252
ROLLING_2M_BUSINESS_DAYS = 42
WINDOW_ORDER = ["last 2 months", "last year", "since 2000", "since 2010", "since 2021"]
SUMMARY_METRICS = [
    "Annual. Return (%)",
    "Annual. Vol. (%)",
    "Sharpe Ratio",
    "Sortino Ratio",
    "Hit* Ratio (%)",
    "4% DD** Quantile (%)",
    "SPX Correl. (wkly)",
    "US10Y Corr. (wkly)",
]
DEFAULT_TOLERANCES = {
    "Annual. Return (%)": 0.02,
    "Annual. Vol. (%)": 0.02,
    "Sharpe Ratio": 0.02,
    "Sortino Ratio": 0.02,
    "Hit* Ratio (%)": 0.05,
    "4% DD** Quantile (%)": 0.05,
    "SPX Correl. (wkly)": 0.02,
    "US10Y Corr. (wkly)": 0.02,
}
DEFAULT_REL_TOLERANCE = 0.01  # 1% relative tolerance as fallback
UNSUPPORTED_METRICS = {"4% DD** Quantile (%)"}
SUMMARY_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


@dataclass
class ValidationArtifacts:
    windows_df: pd.DataFrame
    recomputed_df: pd.DataFrame
    summary_df: pd.DataFrame
    comparison_df: pd.DataFrame
    diagnostics_df: pd.DataFrame


def _derive_daily_return(return_series: pd.Series, mode: str) -> pd.Series:
    clean = pd.to_numeric(return_series, errors="coerce")
    if mode == "diff":
        return clean.diff() / 100.0
    if mode == "raw":
        return clean / 100.0
    raise ValueError(f"Unsupported return derivation mode: {mode}")


def _normalize_columns(df: pd.DataFrame, mapping: dict[str, list[str]], name: str) -> pd.DataFrame:
    result = df.copy()
    for target, candidates in mapping.items():
        for candidate in candidates:
            if candidate in result.columns:
                if candidate != target:
                    result = result.rename(columns={candidate: target})
                break
        else:
            raise ValueError(f"{name} missing required column '{target}'. Available columns: {list(result.columns)}")
    return result


def _annualized_return(daily_returns: pd.Series) -> float:
    clean = pd.to_numeric(daily_returns, errors="coerce").dropna()
    if clean.empty:
        return np.nan
    return float(clean.mean() * TRADING_DAYS * 100)


def _annualized_return_cagr(daily_returns: pd.Series) -> float:
    clean = pd.to_numeric(daily_returns, errors="coerce").dropna()
    if clean.empty:
        return np.nan
    cumulative = (1 + clean).prod()
    return float((cumulative ** (TRADING_DAYS / len(clean)) - 1) * 100)


def _annualized_vol(daily_returns: pd.Series) -> float:
    clean = pd.to_numeric(daily_returns, errors="coerce").dropna()
    if len(clean) < 2:
        return np.nan
    return float(clean.std(ddof=1) * np.sqrt(TRADING_DAYS) * 100)


def _sharpe_ratio(daily_returns: pd.Series) -> float:
    clean = pd.to_numeric(daily_returns, errors="coerce").dropna()
    if len(clean) < 2:
        return np.nan
    std = clean.std(ddof=1)
    if std == 0 or pd.isna(std):
        return np.nan
    return float(clean.mean() / std * np.sqrt(TRADING_DAYS))


def _sortino_ratio(daily_returns: pd.Series, downside_mode: str = "negative_only") -> float:
    """Compute annualized Sortino ratio.

    downside_mode:
        "negative_only" — std of returns where r < 0 (ddof=1)
        "full_series"   — sqrt(mean(min(r, 0)^2)) using all observations
    """
    clean = pd.to_numeric(daily_returns, errors="coerce").dropna()
    if clean.empty:
        return np.nan

    if downside_mode == "full_series":
        downside_sq = np.minimum(clean, 0.0) ** 2
        downside_std = np.sqrt(downside_sq.mean())
    elif downside_mode == "negative_only":
        downside = clean[clean < 0]
        if len(downside) < 2:
            return np.nan
        downside_std = downside.std(ddof=1)
    else:
        raise ValueError(f"Unsupported sortino downside_mode: {downside_mode}")

    if downside_std == 0 or pd.isna(downside_std):
        return np.nan
    return float(clean.mean() / downside_std * np.sqrt(TRADING_DAYS))


def _hit_ratio(daily_returns: pd.Series, hit_mode: str = "strict_positive") -> float:
    """Compute hit ratio as percentage of positive-return days.

    hit_mode:
        "strict_positive"  — count r > 0 as hits; zeros are misses
        "non_negative"     — count r >= 0 as hits; zeros are hits
        "exclude_zeros"    — exclude r == 0 from both numerator and denominator
    """
    clean = pd.to_numeric(daily_returns, errors="coerce").dropna()
    if clean.empty:
        return np.nan

    if hit_mode == "strict_positive":
        return float((clean > 0).mean() * 100)
    elif hit_mode == "non_negative":
        return float((clean >= 0).mean() * 100)
    elif hit_mode == "exclude_zeros":
        non_zero = clean[clean != 0]
        if non_zero.empty:
            return np.nan
        return float((non_zero > 0).mean() * 100)
    else:
        raise ValueError(f"Unsupported hit_mode: {hit_mode}")


def _weekly_corr(window_df: pd.DataFrame, driver_col: str, driver_mode: str) -> float:
    if window_df.empty:
        return np.nan
    weekly_model = (
        window_df.set_index("date")["daily_return"]
        .dropna()
        .resample("W-FRI")
        .apply(lambda s: (1 + s).prod() - 1 if len(s) else np.nan)
    )
    weekly_driver_levels = window_df.set_index("date")[driver_col].dropna().resample("W-FRI").last()
    if driver_mode == "pct_change":
        weekly_driver = weekly_driver_levels.pct_change()
    elif driver_mode == "diff":
        weekly_driver = weekly_driver_levels.diff()
    else:
        raise ValueError(f"Unsupported driver mode: {driver_mode}")

    aligned = pd.concat([weekly_model.rename("model"), weekly_driver.rename("driver")], axis=1).dropna()
    if len(aligned) < 2:
        return np.nan
    return float(aligned["model"].corr(aligned["driver"]))


def load_model_validation_inputs(
    signals_csv: str | Path,
    summary_csv: str | Path,
    model_id: Optional[int] = None,
    return_mode: str = "diff",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    signals_df = pd.read_csv(signals_csv)
    signals_df = _normalize_columns(
        signals_df,
        {
            "date": ["date", "Date", "DATE"],
            "model_id": ["ID", "id", "model_id", "Model_ID"],
            "return_level": ["return"],
            "return_ex_carry": ["return_ex_carry"],
            "SPX": ["SPX"],
            "US10Y": ["US10Y"],
        },
        "Signals CSV",
    )
    signals_df["date"] = pd.to_datetime(signals_df["date"])
    signals_df["model_id"] = pd.to_numeric(signals_df["model_id"], errors="raise").astype(int)
    if model_id is not None:
        signals_df = signals_df[signals_df["model_id"] == int(model_id)].copy()
    signals_df = signals_df.sort_values("date").reset_index(drop=True)
    if signals_df.empty:
        raise ValueError("Signals CSV has no rows after filtering")

    for column in ["return_level", "return_ex_carry", "SPX", "US10Y"]:
        signals_df[column] = pd.to_numeric(signals_df[column], errors="coerce")

    signals_df["daily_return"] = _derive_daily_return(signals_df["return_level"], return_mode)
    signals_df["first_valid_return_date"] = signals_df["return_level"].notna().idxmax()
    signals_df["daily_return_method"] = return_mode

    summary_df = pd.read_csv(summary_csv)
    summary_df = _normalize_columns(
        summary_df,
        {"Evaluation Period": ["Evaluation Period"]},
        "Summary CSV",
    )
    summary_df["Evaluation Period"] = summary_df["Evaluation Period"].astype(str).str.strip()
    for column in SUMMARY_METRICS:
        if column in summary_df.columns:
            summary_df[column] = pd.to_numeric(summary_df[column], errors="coerce")
    summary_df = summary_df[summary_df["Evaluation Period"].isin(WINDOW_ORDER)].copy()
    summary_df["period_order"] = summary_df["Evaluation Period"].map({name: i for i, name in enumerate(WINDOW_ORDER)})
    summary_df = summary_df.sort_values("period_order").drop(columns=["period_order"]).reset_index(drop=True)
    return signals_df, summary_df


def compute_validation_windows(signals_df: pd.DataFrame) -> pd.DataFrame:
    usable = signals_df.loc[signals_df["daily_return"].notna(), ["date", "daily_return"]].copy()
    if usable.empty:
        raise ValueError("Signals CSV contains no usable daily returns")

    latest_date = usable["date"].max()
    first_valid_date = usable["date"].min()
    if len(usable) < ROLLING_2M_BUSINESS_DAYS:
        last_2m_start = first_valid_date
        last_2m_obs = len(usable)
    else:
        trailing_2m = usable.tail(ROLLING_2M_BUSINESS_DAYS)
        last_2m_start = trailing_2m["date"].min()
        last_2m_obs = len(trailing_2m)

    # start_exclusive: True means the daily return ON the start date is excluded
    # because diff(return) on that date straddles outside the window.
    # For trailing business-day windows (tail-based), the start is already the
    # first *valid* return so inclusive is correct.
    definitions = [
        ("last 2 months", last_2m_start, latest_date, False, f"trailing {ROLLING_2M_BUSINESS_DAYS} valid business-day returns"),
        ("last year", latest_date - pd.DateOffset(years=1), latest_date, True, "trailing 1 calendar year (start exclusive)"),
        ("since 2000", pd.Timestamp("2000-01-01"), latest_date, True, "from 2000-01-01 (start exclusive)"),
        ("since 2010", pd.Timestamp("2010-01-01"), latest_date, True, "from 2010-01-01 (start exclusive)"),
        ("since 2021", pd.Timestamp("2021-01-01"), latest_date, True, "from 2021-01-01 (start exclusive)"),
    ]

    rows: list[dict[str, Any]] = []
    for label, start_date, end_date, start_exclusive, method in definitions:
        if start_exclusive:
            date_mask = (signals_df["date"] > start_date) & (signals_df["date"] <= end_date)
        else:
            date_mask = (signals_df["date"] >= start_date) & (signals_df["date"] <= end_date)
        mask = date_mask & signals_df["daily_return"].notna()
        rows.append(
            {
                "Evaluation Period": label,
                "start_date": pd.Timestamp(start_date).date().isoformat(),
                "end_date": pd.Timestamp(end_date).date().isoformat(),
                "start_exclusive": start_exclusive,
                "window_method": method,
                "observation_count": int(last_2m_obs if label == "last 2 months" else mask.sum()),
            }
        )
    return pd.DataFrame(rows)


def recompute_summary_metrics(
    signals_df: pd.DataFrame,
    windows_df: pd.DataFrame,
    sortino_mode: str = "negative_only",
    hit_mode: str = "strict_positive",
) -> pd.DataFrame:
    result_rows: list[dict[str, Any]] = []
    return_mode = str(signals_df["daily_return_method"].dropna().iloc[0]) if "daily_return_method" in signals_df.columns else "diff"
    for window in windows_df.to_dict("records"):
        start_date = pd.Timestamp(window["start_date"])
        end_date = pd.Timestamp(window["end_date"])
        start_exclusive = window.get("start_exclusive", False)
        if start_exclusive:
            window_df = signals_df[(signals_df["date"] > start_date) & (signals_df["date"] <= end_date)].copy()
        else:
            window_df = signals_df[(signals_df["date"] >= start_date) & (signals_df["date"] <= end_date)].copy()
        daily_returns = window_df["daily_return"]
        result_rows.append(
            {
                "Evaluation Period": window["Evaluation Period"],
                "Annual. Return (%)": _annualized_return(daily_returns),
                "Annual Return CAGR (%)": _annualized_return_cagr(daily_returns),
                "Annual. Vol. (%)": _annualized_vol(daily_returns),
                "Sharpe Ratio": _sharpe_ratio(daily_returns),
                "Sortino Ratio": _sortino_ratio(daily_returns, downside_mode=sortino_mode),
                "Hit* Ratio (%)": _hit_ratio(daily_returns, hit_mode=hit_mode),
                "4% DD** Quantile (%)": np.nan,
                "SPX Correl. (wkly)": _weekly_corr(window_df, "SPX", "pct_change"),
                "US10Y Corr. (wkly)": _weekly_corr(window_df, "US10Y", "diff"),
                "Method Note": (
                    f"Daily returns derived as {return_mode}(return). "
                    f"Sortino downside={sortino_mode}. Hit ratio={hit_mode}. "
                    f"Weekly model returns compounded to W-FRI; SPX pct_change, US10Y diff."
                ),
            }
        )
    result_df = pd.DataFrame(result_rows)
    result_df["period_order"] = result_df["Evaluation Period"].map({name: i for i, name in enumerate(WINDOW_ORDER)})
    return result_df.sort_values("period_order").drop(columns=["period_order"]).reset_index(drop=True)


def compare_to_summary(
    recomputed_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    tolerances: Optional[dict[str, float]] = None,
    rel_tolerance: float = DEFAULT_REL_TOLERANCE,
) -> pd.DataFrame:
    """Compare recomputed metrics to source summary.

    Pass/fail uses hybrid tolerance: a metric passes if EITHER the absolute
    delta is within the per-metric tolerance OR the relative delta is within
    ``rel_tolerance``.  This prevents large annualized numbers (e.g. short-window
    returns) from failing on a tight absolute threshold while keeping small
    near-zero metrics honest.
    """
    tolerance_map = {**DEFAULT_TOLERANCES, **(tolerances or {})}
    comparison_rows: list[dict[str, Any]] = []
    summary_by_period = summary_df.set_index("Evaluation Period").to_dict("index")
    recomputed_by_period = recomputed_df.set_index("Evaluation Period").to_dict("index")

    for period in WINDOW_ORDER:
        source_period = summary_by_period.get(period, {})
        recomputed_period = recomputed_by_period.get(period, {})
        for metric in SUMMARY_METRICS:
            source_value = source_period.get(metric, np.nan)
            recomputed_value = recomputed_period.get(metric, np.nan)
            abs_tol = tolerance_map[metric]

            if metric in UNSUPPORTED_METRICS:
                status = "unsupported"
                passed = False
                abs_delta = np.nan
                rel_delta = np.nan
                note = "Not implemented in v1 validator."
            elif pd.isna(source_value):
                status = "missing_source"
                passed = False
                abs_delta = np.nan
                rel_delta = np.nan
                note = "Source summary metric is missing."
            elif pd.isna(recomputed_value):
                status = "missing_recomputed"
                passed = False
                abs_delta = np.nan
                rel_delta = np.nan
                note = "Recomputed metric is missing or insufficient history."
            else:
                abs_delta = abs(float(recomputed_value) - float(source_value))
                rel_delta = np.nan if float(source_value) == 0 else abs_delta / abs(float(source_value))
                abs_pass = abs_delta <= abs_tol
                rel_pass = (not pd.isna(rel_delta)) and (rel_delta <= rel_tolerance)
                passed = abs_pass or rel_pass
                status = "pass" if passed else "fail"
                note = ""

            comparison_rows.append(
                {
                    "Evaluation Period": period,
                    "Metric": metric,
                    "Source Value": source_value,
                    "Recomputed Value": recomputed_value,
                    "Abs Delta": abs_delta,
                    "Rel Delta": rel_delta,
                    "Abs Tolerance": abs_tol,
                    "Rel Tolerance": rel_tolerance,
                    "Status": status,
                    "Passed": passed,
                    "Note": note,
                }
            )

    comparison_df = pd.DataFrame(comparison_rows)
    comparison_df["period_order"] = comparison_df["Evaluation Period"].map({name: i for i, name in enumerate(WINDOW_ORDER)})
    comparison_df["metric_order"] = comparison_df["Metric"].map({name: i for i, name in enumerate(SUMMARY_METRICS)})
    comparison_df = comparison_df.sort_values(["period_order", "metric_order"]).drop(columns=["period_order", "metric_order"]).reset_index(drop=True)
    return comparison_df


def build_diagnostics(comparison_df: pd.DataFrame, top_n: int = 10) -> pd.DataFrame:
    mismatch_df = comparison_df[comparison_df["Status"] == "fail"].copy()
    mismatch_df = mismatch_df.sort_values("Abs Delta", ascending=False).head(top_n)
    unsupported_df = comparison_df[comparison_df["Status"] == "unsupported"].copy()
    unsupported_df = unsupported_df.head(top_n)
    if mismatch_df.empty and unsupported_df.empty:
        return pd.DataFrame([{"Diagnostic": "All supported metrics validated within tolerance."}])

    rows: list[dict[str, Any]] = []
    for record in mismatch_df.to_dict("records"):
        rows.append(
            {
                "Diagnostic": (
                    f"{record['Evaluation Period']} | {record['Metric']} | "
                    f"source={record['Source Value']:.4f} | recomputed={record['Recomputed Value']:.4f} | "
                    f"abs_delta={record['Abs Delta']:.4f} > abs_tol={record['Abs Tolerance']:.4f}"
                )
            }
        )
    for record in unsupported_df.to_dict("records"):
        rows.append({"Diagnostic": f"{record['Evaluation Period']} | {record['Metric']} | {record['Note']}"})
    return pd.DataFrame(rows)


def build_validation_artifacts(
    signals_csv: str | Path,
    summary_csv: str | Path,
    model_id: Optional[int] = None,
    tolerances: Optional[dict[str, float]] = None,
    rel_tolerance: float = DEFAULT_REL_TOLERANCE,
    return_mode: str = "diff",
    sortino_mode: str = "negative_only",
    hit_mode: str = "strict_positive",
) -> ValidationArtifacts:
    signals_df, summary_df = load_model_validation_inputs(signals_csv, summary_csv, model_id=model_id, return_mode=return_mode)
    windows_df = compute_validation_windows(signals_df)
    recomputed_df = recompute_summary_metrics(signals_df, windows_df, sortino_mode=sortino_mode, hit_mode=hit_mode)
    comparison_df = compare_to_summary(recomputed_df, summary_df, tolerances=tolerances, rel_tolerance=rel_tolerance)
    diagnostics_df = build_diagnostics(comparison_df)
    return ValidationArtifacts(
        windows_df=windows_df,
        recomputed_df=recomputed_df,
        summary_df=summary_df,
        comparison_df=comparison_df,
        diagnostics_df=diagnostics_df,
    )


def _style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = SUMMARY_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _autofit(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 12), 48)


def _write_df_sheet(workbook: Workbook, title: str, df: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(df.columns.tolist())
    for row in df.itertuples(index=False, name=None):
        worksheet.append(list(row))
    _style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autofit(worksheet)


def write_validation_report(output_path: str | Path, artifacts: ValidationArtifacts) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_df_sheet(workbook, "Window Definitions", artifacts.windows_df)
    _write_df_sheet(workbook, "Recomputed Metrics", artifacts.recomputed_df)
    _write_df_sheet(workbook, "Source Summary", artifacts.summary_df)
    _write_df_sheet(workbook, "Comparison", artifacts.comparison_df)
    _write_df_sheet(workbook, "Diagnostics", artifacts.diagnostics_df)

    summary_sheet = workbook["Diagnostics"]
    for row in range(2, summary_sheet.max_row + 1):
        summary_sheet.cell(row=row, column=1).fill = SECTION_FILL
        summary_sheet.cell(row=row, column=1).font = Font(bold=True)

    workbook.save(output_path)
    return output_path


def write_comparison_csv(output_path: str | Path, comparison_df: pd.DataFrame) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    comparison_df.to_csv(output_path, index=False)
    return output_path
